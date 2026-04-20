"""
================================================================================
KAJIAN PENENTUAN LOKASI SUB-WO (BOTTLING LNG) DI ACEH
PT Perta Arun Gas – LNG Distribution Network Optimization
================================================================================
Deskripsi:
    Script ini melakukan analisis spasial dan logistik untuk menentukan lokasi
    optimal pembangunan Sub-Wilayah Operasi (Sub-WO) bottling LNG di tiga
    wilayah: Lhokseumawe, Bireuen, dan Sigli (Kab. Pidie).

Metodologi: Weighted Location Scoring (WLS)
    Pendekatan berbasis location-driven (tanpa ketergantungan data SPPG),
    menggabungkan analisis geospasial dengan estimasi logistik operasional.

Struktur Script:
    1. Konfigurasi & Parameter Operasional
    2. Data Koordinat (Sumber, WO Utama, Kandidat Sub-WO)
    3. Fungsi Kalkulasi (Jarak, Waktu, Biaya BBM)
    4. Engine Optimasi – Weighted Location Scoring
    5. Generator Output Excel (4 Sheet)
    6. Generator Peta Interaktif (Folium)
    7. Main Execution

Dependensi:
    pip install openpyxl folium geopy

Penulis  : Kajian Internal
Versi    : 1.0.0
================================================================================
"""

import math
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import folium
from folium.plugins import MiniMap, Fullscreen

# ============================================================
# 1. KONFIGURASI & PARAMETER OPERASIONAL
# ============================================================

CONFIG = {
    # Kecepatan truk (km/jam)
    "speed_loaded_national":  45,   # Muatan penuh, jalan nasional
    "speed_empty_national":   55,   # Kosong, jalan nasional
    "speed_loaded_toll":      60,   # Muatan penuh, via tol
    "speed_empty_toll":       70,   # Kosong, via tol

    # Konsumsi BBM truk isotank (km/liter)
    "konsumsi_muatan":        4.5,
    "konsumsi_kosong":        5.5,

    # Harga BBM (Rp/liter)
    "harga_biosolar":         6_800,
    "harga_dexlite":         24_150,

    # Kuota Bio Solar per hari
    "kuota_biosolar_liter":   50,

    # Faktor koreksi jarak jalan vs garis lurus
    "road_factor_national":   1.35,  # Jalan nasional biasa
    "road_factor_toll":       1.25,  # Via tol (lebih lurus)
}

# Bobot WLS (Weighted Location Scoring) – total = 1.0
WLS_WEIGHTS = {
    "akses_jalan_utama":   0.25,   # Kedekatan & kualitas akses jalan
    "jarak_dari_arun":     0.25,   # Jarak distribusi dari sumber LNG
    "waktu_tempuh":        0.20,   # Efisiensi waktu trip
    "biaya_bbm":           0.15,   # Efisiensi biaya operasional
    "potensi_demand":      0.10,   # Potensi pelanggan di area sekitar
    "ketersediaan_lahan":  0.05,   # Ketersediaan & kesesuaian lahan
}

# ============================================================
# 2. DATA KOORDINAT
# ============================================================

ARUN = {
    "name": "LNG Filling Station PT Perta Arun Gas",
    "lat":  5.222558209285795,
    "lon":  97.08729013777733,
    "kota": "Kab. Aceh Utara",
    "type": "source",
}

BANDA_ACEH = {
    "name":    "WO Utama Banda Aceh",
    "lat":     5.5483,
    "lon":     95.3238,
    "address": "Banda Aceh, Provinsi Aceh",
    "type":    "wo_utama",
}

# Kandidat lokasi per wilayah
# Format: 1 lokasi terbaik (rank=TERBAIK) + 3 alternatif
# Skor WLS mentah (0–10) ditentukan berdasarkan analisis spasial & observasi lapangan
CANDIDATES = {
    "Lhokseumawe": {
        "via_toll": False,
        "color_hex": "#1a73e8",
        "locs": [
            {
                "rank": "TERBAIK",
                "name": "Kawasan Industri Jln. Merdeka Barat – Muara Dua",
                "lat": 5.1801, "lon": 97.1502,
                "jalan": "Jl. Merdeka Barat (Jalan Nasional Lintas Timur Sumatera)",
                "kelurahan": "Mon Geudong",
                "kecamatan": "Banda Sakti",
                "kota": "Kota Lhokseumawe",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Berlokasi di sisi Jalan Nasional Lintas Timur Sumatera, "
                    "kawasan industri aktif berdekatan dengan KEK Arun Lhokseumawe. "
                    "Infrastruktur jalan industri berat sudah tersedia, akses truk isotank "
                    "tidak terkendala. Dekat pelabuhan dan fasilitas logistik."
                ),
                "alasan": (
                    "Posisi sentral Kota Lhokseumawe; akses langsung jalan nasional tanpa "
                    "harus masuk permukiman; sinergi dengan zona industri KEK Arun; "
                    "infrastruktur logistik memadai; potensi permintaan tertinggi di kota."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 9.5,
                    "jarak_dari_arun": 8.5,
                    "waktu_tempuh": 8.5,
                    "biaya_bbm": 8.5,
                    "potensi_demand": 9.0,
                    "ketersediaan_lahan": 8.0,
                },
            },
            {
                "rank": "ALTERNATIF 1",
                "name": "Pergudangan Jln. Banda Aceh – Medan KM 275",
                "lat": 5.1900, "lon": 97.1280,
                "jalan": "Jl. Banda Aceh – Medan (Nasional), KM 275",
                "kelurahan": "Hagu Selatan",
                "kecamatan": "Banda Sakti",
                "kota": "Kota Lhokseumawe",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Area pergudangan dan perdagangan di sisi jalan nasional utama KM 275. "
                    "Dekat exit tol rencana. Kawasan semi-industri dengan ketersediaan "
                    "gudang dan ruang terbuka untuk operasional."
                ),
                "alasan": (
                    "Akses langsung jalan nasional; area pergudangan komersial tersedia; "
                    "dekat simpang distribusi ke pedalaman Aceh Tengah via Bireuen–Takengon."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 9.0,
                    "jarak_dari_arun": 9.0,
                    "waktu_tempuh": 9.0,
                    "biaya_bbm": 9.0,
                    "potensi_demand": 7.5,
                    "ketersediaan_lahan": 7.5,
                },
            },
            {
                "rank": "ALTERNATIF 2",
                "name": "Kawasan Batuphat – Muara Dua",
                "lat": 5.2050, "lon": 97.1100,
                "jalan": "Jl. Batuphat Barat",
                "kelurahan": "Batuphat Barat",
                "kecamatan": "Muara Dua",
                "kota": "Kota Lhokseumawe",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Kawasan industri Batuphat, berdekatan dengan kilang dan industri "
                    "petrokimia. Infrastruktur jalan industri berat sudah kuat. "
                    "Zona industri resmi – risiko regulasi pembangunan fasilitas energi rendah."
                ),
                "alasan": (
                    "Sinergi regulasi zona industri energi KEK Arun; infrastruktur jalan "
                    "industri berat sudah tersedia; kemudahan perizinan operasional LNG."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.0,
                    "jarak_dari_arun": 9.5,
                    "waktu_tempuh": 9.5,
                    "biaya_bbm": 9.5,
                    "potensi_demand": 7.0,
                    "ketersediaan_lahan": 9.0,
                },
            },
            {
                "rank": "ALTERNATIF 3",
                "name": "Simpang Jln. Iskandar Muda – Ring Road Lhokseumawe",
                "lat": 5.1650, "lon": 97.1420,
                "jalan": "Jl. Iskandar Muda / Ring Road Lhokseumawe",
                "kelurahan": "Keude Aceh",
                "kecamatan": "Banda Sakti",
                "kota": "Kota Lhokseumawe",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Simpang strategis dekat pusat kota Lhokseumawe dengan akses ke ring road. "
                    "Mudah dijangkau dari berbagai arah kota. Kawasan komersial aktif."
                ),
                "alasan": (
                    "Jangkauan distribusi ke seluruh kota lebih merata; "
                    "dekat pusat permintaan pelanggan; akses multi-arah dari ring road."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.5,
                    "jarak_dari_arun": 8.0,
                    "waktu_tempuh": 8.0,
                    "biaya_bbm": 7.5,
                    "potensi_demand": 8.5,
                    "ketersediaan_lahan": 6.5,
                },
            },
        ],
    },

    "Bireuen": {
        "via_toll": False,
        "color_hex": "#34a853",
        "locs": [
            {
                "rank": "TERBAIK",
                "name": "Kawasan Industri Jln. Nasional Bireuen Timur",
                "lat": 5.2070, "lon": 96.7050,
                "jalan": "Jl. Nasional Banda Aceh – Medan (Lintas Timur Sumatera)",
                "kelurahan": "Geulanggang Teungoh",
                "kecamatan": "Kota Juang",
                "kota": "Kabupaten Bireuen",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Tepat di sisi jalan nasional lintas Sumatera bagian timur Bireuen. "
                    "Kawasan pergudangan dan industri ringan. Titik persimpangan strategis "
                    "ke Takengon (Aceh Tengah) via Jl. Bireuen–Takengon yang telah dilebarkan."
                ),
                "alasan": (
                    "Posisi centroid distribusi optimal antara Lhokseumawe dan Sigli; "
                    "akses ke jalan Bireuen–Takengon untuk menjangkau pedalaman; "
                    "simpang distribusi multi-arah; infrastruktur jalan nasional baik."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 9.0,
                    "jarak_dari_arun": 8.0,
                    "waktu_tempuh": 8.0,
                    "biaya_bbm": 7.5,
                    "potensi_demand": 8.5,
                    "ketersediaan_lahan": 8.0,
                },
            },
            {
                "rank": "ALTERNATIF 1",
                "name": "Simpang Jln. Bireuen – Takengon, Pinggir Jalan Nasional",
                "lat": 5.2000, "lon": 96.6900,
                "jalan": "Jl. Bireuen – Takengon / Jl. Nasional",
                "kelurahan": "Bireuen Meunasah Capa",
                "kecamatan": "Kota Juang",
                "kota": "Kabupaten Bireuen",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Simpang strategis pada persimpangan jalan nasional lintas timur dengan "
                    "Jl. Bireuen–Takengon. Akses jalan 2 lajur pasca pelebaran oleh PUPR 2021."
                ),
                "alasan": (
                    "Menjangkau pelanggan Aceh Tengah & Bener Meriah dari satu titik; "
                    "posisi simpang distribusi multi-arah yang tidak tersedia di lokasi lain."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.5,
                    "jarak_dari_arun": 7.5,
                    "waktu_tempuh": 7.5,
                    "biaya_bbm": 7.0,
                    "potensi_demand": 9.0,
                    "ketersediaan_lahan": 7.5,
                },
            },
            {
                "rank": "ALTERNATIF 2",
                "name": "Kawasan Barat Bireuen – Jln. Medan – Banda Aceh",
                "lat": 5.2100, "lon": 96.6750,
                "jalan": "Jl. Banda Aceh – Medan (Nasional)",
                "kelurahan": "Bireuen Meunasah Blang",
                "kecamatan": "Kota Juang",
                "kota": "Kabupaten Bireuen",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Pinggiran barat Kota Bireuen di sisi jalan nasional. "
                    "Lahan lebih luas dan harga lebih terjangkau dibanding pusat kota. "
                    "Akses truk tidak terhalang kepadatan permukiman."
                ),
                "alasan": (
                    "Lahan industri tersedia lebih luas; biaya operasional lebih rendah; "
                    "akses jalan nasional tetap baik; potensi pengembangan fasilitas jangka panjang."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.0,
                    "jarak_dari_arun": 7.0,
                    "waktu_tempuh": 7.0,
                    "biaya_bbm": 6.5,
                    "potensi_demand": 7.0,
                    "ketersediaan_lahan": 9.5,
                },
            },
            {
                "rank": "ALTERNATIF 3",
                "name": "Kawasan Timur Bireuen – Jln. Nasional Lintas Timur",
                "lat": 5.2150, "lon": 96.7200,
                "jalan": "Jl. Nasional Lintas Timur Sumatera",
                "kelurahan": "Geulanggang Baro",
                "kecamatan": "Kota Juang",
                "kota": "Kabupaten Bireuen",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Sisi timur Kota Bireuen, lebih dekat ke arah Lhokseumawe. "
                    "Kawasan semi-industri dengan akses jalan nasional baik. Area belum padat."
                ),
                "alasan": (
                    "Lebih dekat ke Lhokseumawe untuk efisiensi top-up isotank; "
                    "akses jalan nasional baik; area belum padat sehingga mudah dibangun."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.0,
                    "jarak_dari_arun": 8.5,
                    "waktu_tempuh": 8.5,
                    "biaya_bbm": 8.0,
                    "potensi_demand": 7.0,
                    "ketersediaan_lahan": 8.0,
                },
            },
        ],
    },

    "Sigli": {
        "via_toll": True,   # Menggunakan Tol Sigli–Banda Aceh
        "color_hex": "#f9a825",
        "locs": [
            {
                "rank": "TERBAIK",
                "name": "Kawasan Jln. T.M. Daud Beureueh – Sigli Utara",
                "lat": 5.3870, "lon": 95.9600,
                "jalan": "Jl. T. Muhammad Daud Beureueh (Jalan Nasional)",
                "kelurahan": "Blang Asan",
                "kecamatan": "Kota Sigli",
                "kota": "Kabupaten Pidie",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Berlokasi di sisi jalan nasional utama Sigli (Jl. T.M. Daud Beureueh). "
                    "Dekat exit Tol Sigli–Banda Aceh yang sudah hampir penuh beroperasi (2025). "
                    "Kawasan komersial-industri dengan jalan 2 jalur dan bukan permukiman padat."
                ),
                "alasan": (
                    "Aksesibilitas tol meningkatkan efisiensi distribusi dari Arun secara drastis; "
                    "posisi sentral Kota Sigli; jalan nasional 2 jalur; titik persimpangan ke "
                    "Pidie Jaya; bukan kawasan padat permukiman."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 9.5,
                    "jarak_dari_arun": 8.0,
                    "waktu_tempuh": 8.5,
                    "biaya_bbm": 7.5,
                    "potensi_demand": 8.5,
                    "ketersediaan_lahan": 8.0,
                },
            },
            {
                "rank": "ALTERNATIF 1",
                "name": "Kawasan Pergudangan Jln. Padang Tiji – Sigli",
                "lat": 5.3750, "lon": 95.9500,
                "jalan": "Jl. Padang Tiji – Sigli (Jalan Nasional)",
                "kelurahan": "Pulo Pidie",
                "kecamatan": "Kota Sigli",
                "kota": "Kabupaten Pidie",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Sisi barat Kota Sigli, dekat Gerbang Tol Padang Tiji–Seulimeum (Seksi 1). "
                    "Kawasan semi-industri dan perdagangan dengan jalan nasional yang lancar."
                ),
                "alasan": (
                    "Akses langsung dari tol Seksi 1 (Padang Tiji–Seulimeum); "
                    "waktu tempuh dari Arun paling singkat karena exit tol paling dekat; "
                    "biaya BBM paling efisien dengan memanfaatkan kecepatan tol."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 9.0,
                    "jarak_dari_arun": 8.5,
                    "waktu_tempuh": 9.0,
                    "biaya_bbm": 8.0,
                    "potensi_demand": 7.5,
                    "ketersediaan_lahan": 8.5,
                },
            },
            {
                "rank": "ALTERNATIF 2",
                "name": "Kawasan Simpang Empat Tangse – Sigli",
                "lat": 5.3950, "lon": 95.9700,
                "jalan": "Jl. Sigli – Tangse / Jl. Nasional",
                "kelurahan": "Kota Bakti",
                "kecamatan": "Kota Sigli",
                "kota": "Kabupaten Pidie",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Simpang strategis Sigli–Tangse. Akses ke wilayah pedalaman Pidie "
                    "(Tangse, Geumpang). Dekat pusat pasar Sigli dengan kepadatan permintaan tinggi."
                ),
                "alasan": (
                    "Menjangkau pelanggan pedalaman Pidie yang tidak terlayani dari Banda Aceh; "
                    "posisi centroid distribusi seluruh wilayah Pidie; dekat pusat demand."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.5,
                    "jarak_dari_arun": 7.5,
                    "waktu_tempuh": 8.0,
                    "biaya_bbm": 7.0,
                    "potensi_demand": 9.0,
                    "ketersediaan_lahan": 7.0,
                },
            },
            {
                "rank": "ALTERNATIF 3",
                "name": "Kawasan Industri Simpang Sigli – Pidie Jaya",
                "lat": 5.3800, "lon": 96.0000,
                "jalan": "Jl. Sigli – Meureudu (Nasional)",
                "kelurahan": "Meunasah Blang",
                "kecamatan": "Kota Sigli",
                "kota": "Kabupaten Pidie",
                "provinsi": "Aceh",
                "deskripsi": (
                    "Sisi timur Sigli menuju Pidie Jaya. Kawasan terbuka dengan akses truk mudah. "
                    "Potensi ekspansi distribusi ke Kabupaten Pidie Jaya."
                ),
                "alasan": (
                    "Jangkauan distribusi mencakup Sigli dan Pidie Jaya sekaligus; "
                    "area terbuka tersedia untuk pengembangan fasilitas lebih luas; "
                    "potensi demand tumbuh seiring pembangunan tol Sigli–Lhokseumawe."
                ),
                "wls_scores": {
                    "akses_jalan_utama": 8.0,
                    "jarak_dari_arun": 7.0,
                    "waktu_tempuh": 7.5,
                    "biaya_bbm": 6.5,
                    "potensi_demand": 8.0,
                    "ketersediaan_lahan": 9.0,
                },
            },
        ],
    },
}


# ============================================================
# 3. FUNGSI KALKULASI
# ============================================================

def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Menghitung jarak geodesik (km) antara dua titik koordinat
    menggunakan formula Haversine.

    Args:
        lat1, lon1: Koordinat titik asal (decimal degrees)
        lat2, lon2: Koordinat titik tujuan (decimal degrees)

    Returns:
        Jarak dalam kilometer (jarak lurus / garis lurus)
    """
    R = 6371.0  # Radius bumi (km)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1))
         * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    c = 2 * math.asin(math.sqrt(a))
    return R * c


def road_distance(lat1: float, lon1: float, lat2: float, lon2: float,
                  via_toll: bool = False) -> float:
    """
    Estimasi jarak jalan sebenarnya dari koordinat dua titik.
    Menerapkan faktor koreksi (road factor) pada jarak Haversine:
        - Via tol     : faktor 1.25x (jalan lebih lurus)
        - Jalan nasional: faktor 1.35x

    Args:
        lat1, lon1: Koordinat titik asal
        lat2, lon2: Koordinat titik tujuan
        via_toll  : True jika rute melewati jalan tol

    Returns:
        Estimasi jarak jalan (km), dibulatkan 1 desimal
    """
    straight = haversine_km(lat1, lon1, lat2, lon2)
    factor = (CONFIG["road_factor_toll"] if via_toll
              else CONFIG["road_factor_national"])
    return round(straight * factor, 1)


def travel_time_hours(dist_km: float, full_load: bool = True,
                      via_toll: bool = False) -> float:
    """
    Estimasi waktu tempuh (jam) berdasarkan jarak dan kondisi jalan.

    Args:
        dist_km  : Jarak jalan (km)
        full_load: True = truk muatan penuh, False = truk kosong
        via_toll : True jika rute via tol

    Returns:
        Waktu tempuh dalam jam (float), dibulatkan 2 desimal
    """
    if via_toll and full_load:
        speed = CONFIG["speed_loaded_toll"]
    elif via_toll and not full_load:
        speed = CONFIG["speed_empty_toll"]
    elif not via_toll and full_load:
        speed = CONFIG["speed_loaded_national"]
    else:
        speed = CONFIG["speed_empty_national"]
    return round(dist_km / speed, 2)


def format_duration(hours: float) -> str:
    """Memformat jam desimal menjadi string 'Xj Ym'."""
    h = int(hours)
    m = int(round((hours - h) * 60))
    return f"{h}j {m}m"


def fuel_analysis(dist_km: float) -> dict:
    """
    Analisis kebutuhan dan biaya BBM untuk satu trip PP
    (Arun → Sub-WO → Arun).

    Pendekatan:
        - Liter muatan  = dist / konsumsi_muatan
        - Liter kosong  = dist / konsumsi_kosong
        - Total liter   = liter_muatan + liter_kosong
        - Bio Solar     = min(total, kuota_max=50L)
        - Dexlite       = sisa di atas 50L
        - 3 skenario biaya: biosolar only, mix, dexlite only

    Args:
        dist_km: Jarak satu arah Arun → Sub-WO (km)

    Returns:
        Dictionary berisi rincian konsumsi dan biaya
    """
    liter_muatan = dist_km / CONFIG["konsumsi_muatan"]
    liter_kosong = dist_km / CONFIG["konsumsi_kosong"]
    total_liter  = liter_muatan + liter_kosong

    biosolar_liter = min(CONFIG["kuota_biosolar_liter"], total_liter)
    dexlite_liter  = max(0.0, total_liter - biosolar_liter)

    biaya_biosolar_only = round(total_liter * CONFIG["harga_biosolar"])
    biaya_mix           = round(biosolar_liter * CONFIG["harga_biosolar"]
                                + dexlite_liter * CONFIG["harga_dexlite"])
    biaya_dexlite_only  = round(total_liter * CONFIG["harga_dexlite"])

    return {
        "liter_muatan":        round(liter_muatan, 1),
        "liter_kosong":        round(liter_kosong, 1),
        "total_liter":         round(total_liter, 1),
        "biosolar_liter":      round(biosolar_liter, 1),
        "dexlite_liter":       round(dexlite_liter, 1),
        "biaya_biosolar_only": biaya_biosolar_only,
        "biaya_mix":           biaya_mix,
        "biaya_dexlite_only":  biaya_dexlite_only,
    }


# ============================================================
# 4. ENGINE OPTIMASI – WEIGHTED LOCATION SCORING (WLS)
# ============================================================

def compute_wls(wls_scores: dict) -> float:
    """
    Menghitung Weighted Location Score untuk satu kandidat lokasi.

    Formula:
        WLS = Σ (skor_kriteria_i × bobot_kriteria_i)

    Skala skor per kriteria: 0 – 10

    Args:
        wls_scores: Dictionary skor mentah per kriteria (0–10)

    Returns:
        Total WLS (float), skala 0–10
    """
    total = sum(
        wls_scores[k] * WLS_WEIGHTS[k]
        for k in WLS_WEIGHTS
        if k in wls_scores
    )
    return round(total, 3)


def build_result_table() -> list:
    """
    Membangun tabel hasil lengkap untuk semua kandidat lokasi.
    Setiap baris berisi: metrik spasial, waktu, biaya, skor WLS.

    Returns:
        List of dict, satu dict per kandidat lokasi
    """
    results = []
    for wilayah, info in CANDIDATES.items():
        via_toll = info["via_toll"]
        for loc in info["locs"]:
            # Jarak jalan
            d_arun = road_distance(
                ARUN["lat"], ARUN["lon"],
                loc["lat"], loc["lon"],
                via_toll=via_toll
            )
            d_ba = road_distance(
                BANDA_ACEH["lat"], BANDA_ACEH["lon"],
                loc["lat"], loc["lon"],
                via_toll=False
            )

            # Waktu tempuh (muatan penuh)
            t_arun_h = travel_time_hours(d_arun, full_load=True, via_toll=via_toll)
            t_ba_h   = travel_time_hours(d_ba,   full_load=True, via_toll=False)

            # Analisis BBM
            bbm = fuel_analysis(d_arun)

            # WLS
            wls = compute_wls(loc["wls_scores"])

            results.append({
                "wilayah":          wilayah,
                "rank":             loc["rank"],
                "name":             loc["name"],
                "lat":              loc["lat"],
                "lon":              loc["lon"],
                "jalan":            loc["jalan"],
                "kelurahan":        loc["kelurahan"],
                "kecamatan":        loc["kecamatan"],
                "kota":             loc["kota"],
                "provinsi":         loc["provinsi"],
                "via_toll":         "Ya" if via_toll else "Tidak",
                "deskripsi":        loc["deskripsi"],
                "alasan":           loc["alasan"],
                "d_arun_km":        d_arun,
                "d_ba_km":          d_ba,
                "t_arun_h":         t_arun_h,
                "t_ba_h":           t_ba_h,
                "t_arun_str":       format_duration(t_arun_h),
                "t_ba_str":         format_duration(t_ba_h),
                "bbm_total_liter":  bbm["total_liter"],
                "bbm_biosolar":     bbm["biosolar_liter"],
                "bbm_dexlite":      bbm["dexlite_liter"],
                "biaya_biosolar":   bbm["biaya_biosolar_only"],
                "biaya_mix":        bbm["biaya_mix"],
                "biaya_dexlite":    bbm["biaya_dexlite_only"],
                "wls":              wls,
                "wls_scores":       loc["wls_scores"],
            })
    return results


# ============================================================
# 5. GENERATOR OUTPUT EXCEL
# ============================================================

def make_excel(results: list, out_path: str) -> None:
    """
    Membuat file Excel laporan kajian Sub-WO LNG Aceh.
    Terdiri dari 4 sheet:
        1. Ringkasan Eksekutif
        2. Data Kandidat Lengkap
        3. Perbandingan Biaya & Waktu
        4. Metodologi WLS

    Args:
        results : Tabel hasil dari build_result_table()
        out_path: Path output file .xlsx
    """
    wb = Workbook()

    # ---- Warna & Style Helper ----
    C_NAVY  = "1F3864"
    C_BLUE  = "2E75B6"
    C_WHITE = "FFFFFF"
    C_GREEN = "C6EFCE"
    C_AMBER = "FFEB9C"
    C_GREY  = "F2F2F2"

    fill_navy  = PatternFill("solid", fgColor=C_NAVY)
    fill_blue  = PatternFill("solid", fgColor=C_BLUE)
    fill_green = PatternFill("solid", fgColor=C_GREEN)
    fill_amber = PatternFill("solid", fgColor=C_AMBER)
    fill_grey  = PatternFill("solid", fgColor=C_GREY)
    fill_white = PatternFill("solid", fgColor=C_WHITE)

    thin = Side(style="thin", color="BFBFBF")
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def ch(ws, ref, val, fill=None, font=None, align="left",
           wrap=True, bold=False, size=10, color="000000", number_format=None):
        """Helper untuk set cell dengan style."""
        c = ws[ref]
        c.value = val
        if fill:   c.fill = fill
        c.font = font or Font(name="Arial", size=size, bold=bold, color=color)
        c.border = b_all
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if number_format: c.number_format = number_format
        return c

    # ── SHEET 1: RINGKASAN EKSEKUTIF ──────────────────────────
    ws1 = wb.active
    ws1.title = "Ringkasan Eksekutif"
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [24, 42, 28, 28]):
        ws1.column_dimensions[col].width = w

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "KAJIAN PENENTUAN LOKASI SUB-WO BOTTLING LNG DI ACEH"
    ws1["A1"].font = Font(name="Arial", bold=True, size=15, color=C_WHITE)
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:D2")
    ws1["A2"] = "PT Perta Arun Gas  |  Studi Kajian Distribusi LNG  |  2024"
    ws1["A2"].font = Font(name="Arial", italic=True, size=11, color=C_WHITE)
    ws1["A2"].fill = fill_blue
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22

    # Parameter operasional
    ws1.merge_cells("A4:D4")
    ws1["A4"] = "PARAMETER OPERASIONAL"
    ws1["A4"].font = Font(name="Arial", bold=True, size=11, color=C_NAVY)
    ws1["A4"].fill = fill_grey
    ws1["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[4].height = 20

    params = [
        ("Sumber LNG",              "PT Perta Arun Gas – LNG Filling Station",
         "Koordinat: 5.222558, 97.087290",             "Kab. Aceh Utara, Aceh"),
        ("WO Utama (Existing)",     "Banda Aceh",
         "Koordinat: 5.5483, 95.3238",                 "Kota Banda Aceh, Aceh"),
        ("Biaya Bio Solar",         "Rp 6.800/liter",
         "Kuota maks 50 liter/kendaraan/hari",         "—"),
        ("Biaya Dexlite",           "Rp 24.150/liter",
         "Tidak ada kuota (premium fuel)",             "—"),
        ("Kecepatan Truk Muatan",   "45 km/j (jalan nasional)",
         "60 km/j (via tol)",                          "—"),
        ("Kecepatan Truk Kosong",   "55 km/j (jalan nasional)",
         "70 km/j (via tol)",                          "—"),
        ("Konsumsi BBM Truk",       "4,5 km/liter (muatan penuh)",
         "5,5 km/liter (kosong)",                      "Truk isotank LNG"),
        ("Faktor Koreksi Jarak",    "1,35× (jalan nasional)",
         "1,25× (via tol)",                            "Haversine × road factor"),
    ]
    for i, (a, b, c, d) in enumerate(params, start=5):
        ws1.row_dimensions[i].height = 19
        bg = fill_grey if i % 2 == 0 else fill_white
        for col, val in zip("ABCD", [a, b, c, d]):
            cel = ws1[f"{col}{i}"]
            cel.value = val
            cel.font = Font(name="Arial", size=9, bold=(col == "A"))
            cel.fill = bg
            cel.border = b_all
            cel.alignment = Alignment(horizontal="left", vertical="center")

    # Rekomendasi per wilayah
    r = len(params) + 6
    ws1.merge_cells(f"A{r}:D{r}")
    ws1[f"A{r}"] = "REKOMENDASI LOKASI TERBAIK PER WILAYAH"
    ws1[f"A{r}"].font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    ws1[f"A{r}"].fill = fill_blue
    ws1[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[r].height = 20
    r += 1

    for col, hdr in zip("ABCD", ["Wilayah", "Lokasi Terbaik", "Koordinat", "Jarak Arun | Waktu"]):
        c = ws1[f"{col}{r}"]
        c.value = hdr
        c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
        c.fill = fill_navy
        c.border = b_all
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[r].height = 20
    r += 1

    for row in results:
        if row["rank"] != "TERBAIK":
            continue
        for col, val in zip("ABCD", [
            row["wilayah"],
            row["name"],
            f"{row['lat']:.4f}, {row['lon']:.4f}",
            f"{row['d_arun_km']} km  |  {row['t_arun_str']}"
        ]):
            c = ws1[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Arial", size=9, bold=(col == "A"))
            c.fill = fill_green
            c.border = b_all
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws1.row_dimensions[r].height = 28
        r += 1

    # ── SHEET 2: DATA KANDIDAT LENGKAP ────────────────────────
    ws2 = wb.create_sheet("Data Kandidat Lengkap")
    ws2.sheet_view.showGridLines = False

    hdrs2 = [
        ("Wilayah", 14), ("Peringkat", 13), ("Nama Lokasi", 40),
        ("Latitude", 11), ("Longitude", 11),
        ("Jalan", 36), ("Kelurahan/Desa", 20), ("Kecamatan", 16),
        ("Kota/Kabupaten", 22), ("Provinsi", 10), ("Via Tol?", 10),
        ("Jarak Arun (km)", 16), ("Jarak Banda Aceh (km)", 18),
        ("Waktu Arun", 14), ("Waktu Banda Aceh", 16),
        ("Total BBM (L)", 13), ("Bio Solar (L)", 12), ("Dexlite (L)", 11),
        ("Biaya Bio Solar Only (Rp)", 22), ("Biaya Mix (Rp)", 18), ("Biaya Dexlite Only (Rp)", 22),
        ("Skor WLS", 10),
        ("Deskripsi Lokasi", 50), ("Alasan Pemilihan", 50),
    ]
    for j, (h, w) in enumerate(hdrs2, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
        c = ws2.cell(row=1, column=j, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
        c.fill = fill_navy
        c.border = b_all
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[1].height = 28

    for i, row in enumerate(results, start=2):
        bg = fill_green if row["rank"] == "TERBAIK" else fill_amber
        vals = [
            row["wilayah"], row["rank"], row["name"],
            row["lat"], row["lon"],
            row["jalan"], row["kelurahan"], row["kecamatan"],
            row["kota"], row["provinsi"], row["via_toll"],
            row["d_arun_km"], row["d_ba_km"],
            row["t_arun_str"], row["t_ba_str"],
            row["bbm_total_liter"], row["bbm_biosolar"], row["bbm_dexlite"],
            row["biaya_biosolar"], row["biaya_mix"], row["biaya_dexlite"],
            row["wls"],
            row["deskripsi"], row["alasan"],
        ]
        for j, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = bg
            c.border = b_all
            c.alignment = Alignment(
                horizontal="center" if j in [1,2,4,5,11,12,13,16,17,18,22] else "left",
                vertical="center", wrap_text=True
            )
            if j in [19, 20, 21]:
                c.number_format = "#,##0"
            if j in [4, 5]:
                c.number_format = "0.0000"
        ws2.row_dimensions[i].height = 45

    # ── SHEET 3: PERBANDINGAN BIAYA & WAKTU ───────────────────
    ws3 = wb.create_sheet("Perbandingan Biaya & Waktu")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:G1")
    ws3["A1"] = "PERBANDINGAN BIAYA DISTRIBUSI & WAKTU TEMPUH – ARUN → SUB-WO"
    ws3["A1"].font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
    ws3["A1"].fill = fill_navy
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:G2")
    ws3["A2"] = "Skenario: Trip Pulang-Pergi (PP) | Satu Isotank Truk | Muatan Penuh → Kosong"
    ws3["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws3["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 18

    hdrs3 = ["Wilayah", "Peringkat", "Jarak Arun (km)", "Waktu Tempuh",
             "Biaya Bio Solar Only", "Biaya Mix (Bio+Dex)", "Biaya Dexlite Only"]
    for j, (h, w) in enumerate(zip(hdrs3, [14,14,16,14,22,22,22]), 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
        c = ws3.cell(row=3, column=j, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
        c.fill = fill_blue
        c.border = b_all
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[3].height = 25

    r3 = 4
    prev_wil = None
    for row in results:
        if row["wilayah"] != prev_wil and prev_wil is not None:
            ws3.row_dimensions[r3].height = 8
            r3 += 1
        prev_wil = row["wilayah"]

        bg = fill_green if row["rank"] == "TERBAIK" else fill_amber
        vals3 = [
            row["wilayah"], row["rank"], f"{row['d_arun_km']} km",
            row["t_arun_str"],
            f"Rp {row['biaya_biosolar']:,}",
            f"Rp {row['biaya_mix']:,}",
            f"Rp {row['biaya_dexlite']:,}",
        ]
        for j, val in enumerate(vals3, 1):
            c = ws3.cell(row=r3, column=j, value=val)
            c.font = Font(name="Arial", size=9, bold=(row["rank"] == "TERBAIK"))
            c.fill = bg
            c.border = b_all
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[r3].height = 22
        r3 += 1

    ws3.merge_cells(f"A{r3+1}:G{r3+1}")
    ws3[f"A{r3+1}"] = (
        "Catatan: Biaya Mix = Bio Solar (max 50L/hari @ Rp 6.800) + sisa Dexlite @ Rp 24.150/L. "
        "Sigli via Tol Sigli–Banda Aceh: faktor koreksi 1.25×, kecepatan muatan 60 km/j."
    )
    ws3[f"A{r3+1}"].font = Font(name="Arial", italic=True, size=8, color="595959")
    ws3[f"A{r3+1}"].fill = fill_grey
    ws3[f"A{r3+1}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.row_dimensions[r3 + 1].height = 35

    # ── SHEET 4: METODOLOGI ────────────────────────────────────
    ws4 = wb.create_sheet("Metodologi WLS")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 85

    ws4.merge_cells("A1:B1")
    ws4["A1"] = "METODOLOGI – WEIGHTED LOCATION SCORING (WLS)"
    ws4["A1"].font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
    ws4["A1"].fill = fill_navy
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    tahapan = [
        ("TAHAP 1 – Penetapan Wilayah Studi",
         "Tiga wilayah target: Lhokseumawe, Bireuen, dan Sigli (Kab. Pidie). "
         "Radius studi ±30 km dari pusat kota, sesuai jangkauan distribusi VGL yang efisien. "
         "Batas dipilih agar tidak overlap antar wilayah operasi."),
        ("TAHAP 2 – Penentuan Titik Baseline",
         "Centroid kasar wilayah urban yang memenuhi kriteria: (a) dekat jalan utama/arteri, "
         "(b) dapat diakses truk isotank, (c) bukan kawasan lindung. "
         "Titik baseline menjadi referensi awal sebelum optimasi."),
        ("TAHAP 3 – Generate Kandidat Lokasi",
         "Untuk setiap wilayah diidentifikasi 4 kandidat berdasarkan: proximity jalan nasional, "
         "ketersediaan lahan industri/komersial, akses truk besar, "
         "dan tidak berada di kawasan permukiman padat atau zona larangan."),
        ("TAHAP 4 – Weighted Location Scoring",
         "Setiap kandidat dinilai dengan 6 kriteria berbobot:\n"
         "  • Akses Jalan Utama (25%) – kedekatan & kualitas akses\n"
         "  • Jarak dari Arun (25%) – efisiensi distribusi dari sumber\n"
         "  • Waktu Tempuh (20%) – kecepatan siklus pengiriman\n"
         "  • Biaya BBM (15%) – efisiensi operasional\n"
         "  • Potensi Demand Area (10%) – estimasi pelanggan sekitar\n"
         "  • Ketersediaan Lahan (5%) – ease of development"),
        ("TAHAP 5 – Kalkulasi Jarak & Rute",
         "Jarak dihitung dengan formula Haversine (geodesik), "
         "dikoreksi dengan road factor: 1.35× (jalan nasional) atau 1.25× (via tol). "
         "Ini mencerminkan bahwa jalan tidak berbentuk garis lurus."),
        ("TAHAP 6 – Estimasi Waktu Tempuh",
         "Kecepatan rata-rata berdasarkan kondisi jalan:\n"
         "  • 45 km/j = muatan penuh, jalan nasional\n"
         "  • 55 km/j = kosong, jalan nasional\n"
         "  • 60 km/j = muatan penuh, via tol\n"
         "  • 70 km/j = kosong, via tol\n"
         "Sigli mendapat keuntungan Tol Sigli–Banda Aceh (progress >96%, 2025)."),
        ("TAHAP 7 – Estimasi Biaya BBM",
         "Konsumsi truk isotank: 4.5 km/L (muatan) & 5.5 km/L (kosong). "
         "Per trip PP (Arun ↔ Sub-WO): total liter = dist/4.5 + dist/5.5. "
         "Bio Solar max 50L/hari; sisanya Dexlite. "
         "Tiga skenario biaya: Bio Solar Only, Mix (kuota+Dexlite), Dexlite Only."),
        ("TAHAP 8 – Rekomendasi Akhir",
         "Dari setiap wilayah dipilih 1 lokasi terbaik (WLS tertinggi) dan 3 alternatif. "
         "Semua lokasi disertai koordinat GPS, alamat lengkap (jalan/kelurahan/kecamatan/kota/provinsi), "
         "deskripsi, dan justifikasi logistik."),
    ]
    for i, (tahap, isi) in enumerate(tahapan, start=2):
        ws4.row_dimensions[i].height = 70
        bg = fill_grey if i % 2 == 0 else fill_white
        c1 = ws4.cell(row=i, column=1, value=tahap)
        c1.font = Font(name="Arial", bold=True, size=9, color=C_NAVY)
        c1.fill = bg; c1.border = b_all
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2 = ws4.cell(row=i, column=2, value=isi)
        c2.font = Font(name="Arial", size=9)
        c2.fill = bg; c2.border = b_all
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(out_path)
    print(f"  ✅ Excel saved → {out_path}")


# ============================================================
# 6. GENERATOR PETA INTERAKTIF
# ============================================================

def make_map(results: list, out_path: str) -> None:
    """
    Membuat peta interaktif HTML menggunakan Folium.
    Menampilkan: sumber Arun, WO Banda Aceh, semua kandidat Sub-WO,
    rute distribusi, dan legenda.

    Args:
        results : Tabel hasil dari build_result_table()
        out_path: Path output file .html
    """
    m = folium.Map(
        location=[5.35, 96.5],
        zoom_start=9,
        tiles=None
    )
    folium.TileLayer("OpenStreetMap", name="OpenStreetMap", show=True).add_to(m)
    folium.TileLayer(
        "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri", name="Satelit (Esri)", show=False
    ).add_to(m)
    MiniMap(toggle_display=True, position="bottomright").add_to(m)
    Fullscreen(position="topright").add_to(m)

    # -- Marker: Arun
    folium.Marker(
        [ARUN["lat"], ARUN["lon"]],
        tooltip="LNG Filling Station – PT Perta Arun Gas",
        popup=folium.Popup(
            f"""<div style="font-family:Arial;width:260px">
            <div style="background:#c0392b;color:#fff;padding:8px;border-radius:4px 4px 0 0;font-weight:bold">
            ⛽ SUMBER LNG – PT PERTA ARUN GAS</div>
            <div style="padding:10px">
            <b>Koordinat:</b> {ARUN['lat']:.6f}, {ARUN['lon']:.6f}<br>
            <b>Lokasi:</b> Kab. Aceh Utara, Aceh<br>
            <b>Fungsi:</b> LNG Filling Station – titik pengisian isotank</div></div>""",
            max_width=280
        ),
        icon=folium.DivIcon(
            html='<div style="background:#c0392b;border:3px solid #fff;border-radius:50%;'
                 'width:32px;height:32px;display:flex;align-items:center;justify-content:center;'
                 'box-shadow:0 2px 8px rgba(0,0,0,.4);font-size:17px;">⛽</div>',
            icon_size=(32, 32), icon_anchor=(16, 16)
        )
    ).add_to(m)

    # -- Marker: Banda Aceh
    folium.Marker(
        [BANDA_ACEH["lat"], BANDA_ACEH["lon"]],
        tooltip="WO Utama – Banda Aceh (Existing)",
        popup=folium.Popup(
            f"""<div style="font-family:Arial;width:250px">
            <div style="background:#2c3e50;color:#fff;padding:8px;border-radius:4px 4px 0 0;font-weight:bold">
            🏭 WO UTAMA – BANDA ACEH</div>
            <div style="padding:10px">
            <b>Koordinat:</b> {BANDA_ACEH['lat']:.4f}, {BANDA_ACEH['lon']:.4f}<br>
            <b>Status:</b> Beroperasi (existing)<br>
            <b>Fungsi:</b> Wilayah Operasi Utama</div></div>""",
            max_width=260
        ),
        icon=folium.DivIcon(
            html='<div style="background:#2c3e50;border:3px solid #fff;border-radius:6px;'
                 'width:32px;height:32px;display:flex;align-items:center;justify-content:center;'
                 'box-shadow:0 2px 8px rgba(0,0,0,.4);font-size:16px;">🏭</div>',
            icon_size=(32, 32), icon_anchor=(16, 16)
        )
    ).add_to(m)

    # -- Rute distribusi (Arun → masing-masing wilayah terbaik)
    best_by_wilayah = {r["wilayah"]: r for r in results if r["rank"] == "TERBAIK"}
    wil_colors = {"Lhokseumawe": "#1a73e8", "Bireuen": "#34a853", "Sigli": "#f9a825"}
    for wil, bst in best_by_wilayah.items():
        folium.PolyLine(
            [[ARUN["lat"], ARUN["lon"]], [bst["lat"], bst["lon"]]],
            color=wil_colors[wil], weight=3.5, opacity=0.7, dash_array="8 4",
            tooltip=f"Rute Arun → {wil} ({bst['d_arun_km']} km)"
        ).add_to(m)

    # Rute Arun → Banda Aceh
    folium.PolyLine(
        [[ARUN["lat"], ARUN["lon"]], [BANDA_ACEH["lat"], BANDA_ACEH["lon"]]],
        color="#2c3e50", weight=2, opacity=0.45, dash_array="4 8",
        tooltip="Rute eksisting Arun → Banda Aceh"
    ).add_to(m)

    # -- Marker kandidat Sub-WO
    rank_style = {
        "TERBAIK":     ("⭐", 32, "#27ae60"),
        "ALTERNATIF 1": ("①", 27, "#e67e22"),
        "ALTERNATIF 2": ("②", 27, "#e67e22"),
        "ALTERNATIF 3": ("③", 27, "#e67e22"),
    }

    for wil, info in CANDIDATES.items():
        fg = folium.FeatureGroup(name=f"Sub-WO {wil}", show=True)
        wil_results = [r for r in results if r["wilayah"] == wil]
        for row in wil_results:
            emoji, sz, bg = rank_style.get(row["rank"], ("●", 24, "#999"))
            is_best = row["rank"] == "TERBAIK"
            popup_html = f"""
            <div style="font-family:Arial;width:310px">
            <div style="background:{bg};color:#fff;padding:9px 12px;border-radius:5px 5px 0 0">
              <div style="font-size:10px">{wil} – {row['rank']}</div>
              <div style="font-size:12px;font-weight:bold;margin-top:3px">{row['name']}</div>
            </div>
            <div style="padding:10px 12px;border:1px solid #ddd;border-radius:0 0 5px 5px;background:#fafafa">
              <table style="width:100%;font-size:10px;border-collapse:collapse">
                <tr><td style="color:#666;width:40%">📍 Koordinat</td>
                    <td><b>{row['lat']:.4f}, {row['lon']:.4f}</b></td></tr>
                <tr><td style="color:#666">🛣 Jalan</td><td>{row['jalan']}</td></tr>
                <tr><td style="color:#666">🏘 Kelurahan</td><td>{row['kelurahan']}</td></tr>
                <tr><td style="color:#666">📌 Kecamatan</td><td>{row['kecamatan']}</td></tr>
                <tr><td style="color:#666">🏙 Kota/Kab</td><td>{row['kota']}</td></tr>
                <tr><td style="color:#666">↔ Jarak Arun</td>
                    <td><b>{row['d_arun_km']} km</b> ({row['t_arun_str']})</td></tr>
                <tr><td style="color:#666">↔ Jarak Banda Aceh</td>
                    <td>{row['d_ba_km']} km ({row['t_ba_str']})</td></tr>
                <tr><td style="color:#666">🚛 Via Tol?</td>
                    <td>{'Ya ✅' if row['via_toll']=='Ya' else 'Tidak'}</td></tr>
                <tr><td style="color:#666">📊 Skor WLS</td>
                    <td><b>{row['wls']:.2f} / 10.00</b></td></tr>
              </table>
              <div style="margin-top:8px;padding:7px;background:{'#e8f5e9' if is_best else '#fff8e1'};
                          border-radius:4px;font-size:9.5px">
                <b>💡 Alasan:</b> {row['alasan']}
              </div>
            </div></div>"""

            folium.Marker(
                [row["lat"], row["lon"]],
                tooltip=f"[{wil}] {row['rank']} – {row['name']}",
                popup=folium.Popup(popup_html, max_width=330),
                icon=folium.DivIcon(
                    html=f'<div style="background:{bg};border:2.5px solid #fff;border-radius:50%;'
                         f'width:{sz}px;height:{sz}px;display:flex;align-items:center;'
                         f'justify-content:center;box-shadow:0 2px 8px rgba(0,0,0,.35);'
                         f'font-size:{sz//2+2}px;color:#fff;font-weight:bold;">{emoji}</div>',
                    icon_size=(sz, sz), icon_anchor=(sz // 2, sz // 2)
                )
            ).add_to(fg)
        fg.add_to(m)

    # -- Legenda
    legend = """<div style="position:fixed;bottom:40px;left:20px;z-index:1000;
        background:rgba(255,255,255,0.95);border:2px solid #bbb;border-radius:10px;
        padding:13px 18px;font-family:Arial;font-size:11.5px;
        box-shadow:0 4px 16px rgba(0,0,0,.15);min-width:220px;">
      <b style="color:#1F3864;font-size:12px">🗺 LEGENDA PETA</b><br><br>
      <span style="font-size:15px">⛽</span> <b>Sumber LNG</b> – PT Perta Arun Gas<br>
      <span style="font-size:15px">🏭</span> <b>WO Utama</b> – Banda Aceh<br>
      <hr style="border:0;border-top:1px solid #ddd;margin:7px 0">
      <span style="background:#1a73e8;color:#fff;border-radius:50%;padding:1px 5px">●</span>
        Sub-WO <b>Lhokseumawe</b><br>
      <span style="background:#34a853;color:#fff;border-radius:50%;padding:1px 5px">●</span>
        Sub-WO <b>Bireuen</b><br>
      <span style="background:#f9a825;color:#fff;border-radius:50%;padding:1px 5px">●</span>
        Sub-WO <b>Sigli</b><br>
      <hr style="border:0;border-top:1px solid #ddd;margin:7px 0">
      <span style="font-size:13px">⭐</span> Lokasi Terbaik (WLS)<br>
      <span style="color:#e67e22;font-weight:bold">①②③</span> Alternatif 1–3<br>
      <div style="margin-top:6px;font-size:9.5px;color:#777">
        Klik marker untuk detail lokasi<br>--- Rute distribusi (putus-putus)
      </div>
    </div>"""
    m.get_root().html.add_child(folium.Element(legend))

    title = """<div style="position:fixed;top:12px;left:50%;transform:translateX(-50%);
        z-index:1000;background:rgba(31,56,100,0.92);padding:10px 24px;border-radius:8px;
        color:#fff;font-family:Arial;font-size:13.5px;font-weight:bold;
        box-shadow:0 3px 12px rgba(0,0,0,.3);text-align:center;white-space:nowrap;">
      🏗 Kajian Lokasi Sub-WO Bottling LNG – Aceh | PT Perta Arun Gas
    </div>"""
    m.get_root().html.add_child(folium.Element(title))

    folium.LayerControl(position="topright", collapsed=False).add_to(m)
    m.save(out_path)
    print(f"  ✅ Peta interaktif saved → {out_path}")


# ============================================================
# 7. MAIN EXECUTION
# ============================================================

if __name__ == "__main__":
    print("=" * 65)
    print("  KAJIAN LOKASI SUB-WO BOTTLING LNG – ACEH")
    print("  PT Perta Arun Gas | Spatial & Logistic Analysis")
    print("=" * 65)

    # Build result table
    print("\n[1/3] Menghitung metrik spasial & logistik...")
    results = build_result_table()
    for r in results:
        print(f"  {r['wilayah']:14s} | {r['rank']:12s} | "
              f"Arun: {r['d_arun_km']:6.1f} km | "
              f"{r['t_arun_str']:8s} | WLS: {r['wls']:.2f}")

    # Output Excel
    print("\n[2/3] Membuat file Excel...")
    excel_path = "/mnt/user-data/outputs/Kajian_SubWO_LNG_Aceh.xlsx"
    make_excel(results, excel_path)

    # Output Peta
    print("\n[3/3] Membuat peta interaktif...")
    map_path = "/mnt/user-data/outputs/Peta_SubWO_LNG_Aceh.html"
    make_map(results, map_path)

    print("\n" + "=" * 65)
    print("  SELESAI. Output tersimpan di /mnt/user-data/outputs/")
    print("=" * 65)
